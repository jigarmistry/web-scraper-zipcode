import csv
import time
import json
import argparse
import requests
import openpyxl as px
from bs4 import BeautifulSoup
from selenium import webdriver
from operator import itemgetter

parser = argparse.ArgumentParser(description='Zipcode and radius')
parser.add_argument('-r', action='store', dest='radius',
                    help='Radius in miles')
parser.add_argument('-z', action='store', dest='zip_code',
                    help='Zipcode')


browser = webdriver.PhantomJS("/usr/local/lib/phantomjs/bin/phantomjs", service_args=[
                              '--ssl-protocol=any', '--ignore-ssl-errors=true', '--load-images=no'])

browser.set_window_size(1600, 900)

url_zipcode = "https://www.freemaptools.com/find-zip-codes-inside-radius.htm"
url_grades = "http://education.ohio.gov/getattachment/Topics/Data/Report-Card-Resources/DISTRICT-GRADES.xlsx"


def get_zipcodes(miles, zip_code):

    browser.get(url_zipcode)

    radiusElem = browser.find_element_by_id('tb_radius_miles')
    radiusElem.send_keys(miles)
    zipCodeElem = browser.find_element_by_id('goto')
    zipCodeElem.send_keys(zip_code)
    drawRadiusBtn = browser.find_element_by_name('Go')
    drawRadiusBtn.click()
    time.sleep(5)

    html_data = browser.find_element_by_id("tb_output").get_attribute('value')
    zip_list = html_data.split(',')
    return zip_list


def get_district_grades(zip_list):

    download_data = requests.get(url_grades).content
    xls_file = open('district_grades_input.xlsx', 'wb')
    xls_file.write(download_data)
    xls_file.close()

    grades = []
    a = 0
    W = px.load_workbook('district_grades_input.xlsx')
    p = W.get_sheet_by_name(name='DISTRICT')
    for row in p.iter_rows():
        in_list = []
        if a != 0:
            for k in row:
                in_list.append(k.internal_value)
            grades.append(in_list)
        a = a + 1

    grades.sort(key=lambda row: row[10])

    output_grades = []
    for i in grades:
        print id
    wb = px.Workbook()
    dest_filename = 'empty_book.xlsx'
    ws1 = wb.active
    ws1.title = "range names"
    ws1.append([1, 2, 3])
    wb.save(filename=dest_filename)

if __name__ == "__main__":

    options = parser.parse_args()
    if options.radius and options.zip_code:
        zip_list = get_zipcodes(options.radius, options.zip_code)
        if zip_list != []:
            get_district_grades(zip_list)
        else:
            print ("No Zipcodes Founds")
    else:
        print("Please provide the zipcode/radius")
        parser.print_help()


#free[u'44802', u'45890', u'43316', u'44853', u'44804', u'44830', u'44844', u'44809', u'45867', u'45839', u'43359', u'44817', u'45889', u'45840', u'44883', u'45814', u'43467', u'43457', u'44845', u'43330', u'44882', u'44841', u'43351', u'44815', u'43437', u'45872', u'43466', u'45897', u'43407', u'43413', u'45858', u'45816', u'45841', u'45881', u'44861', u'45843', u'43435', u'43406', u'43529', u'43451', u'45836', u'44860', u'45868', u'43462', u'44818', u'43323', u'44849', u'45835', u'44836', u'43431', u'43450', u'43511', u'43403', u'44867', u'44825', u'45817', u'43337', u'43420', u'45810', u'45877', u'43541', u'45815', u'43516', u'43442', u'43402', u'44820', u'44807', u'43569', u'43469', u'43443', u'43326', u'43414', u'43565', u'45859', u'44854', u'43332', u'44881']
